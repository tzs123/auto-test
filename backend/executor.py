"""测试执行器：隔离结果目录、流式日志、结果解析、报告生成、飞书通知。

负载均衡下多 worker 并发执行时，每个任务使用独立 results/report 目录，互不干扰。
"""
import os
import sys
import json
import glob
import time
import uuid
import shutil
import subprocess
import yaml

# 支持直接执行（python -m backend.executor）与包内导入
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend import db, log_store, settings, feishu


def _now():
    return time.strftime("%Y-%m-%d %H:%M:%S")


def _task_dirs(task_id: str) -> tuple:
    base = os.path.join(settings.REPORT_DIR, "tasks", task_id)
    return base, os.path.join(base, "results"), os.path.join(base, "report")


def stop_task(task_id: str) -> dict:
    """停止正在运行的任务，终止子进程组。"""
    proc = _running_procs.get(task_id)
    if not proc:
        return {"ok": False, "msg": "任务不在运行中或已结束"}

    import signal as _signal
    try:
        os.killpg(os.getpgid(proc.pid), _signal.SIGTERM)
        log_store.write(task_id, "⛔ 用户手动停止任务，已发送 SIGTERM\n")
    except ProcessLookupError:
        pass
    except Exception as e:
        log_store.write(task_id, f"⛔ 停止任务异常: {e}\n")

    _update_task(task_id, status="stopped", finished_at=_now())
    return {"ok": True, "msg": "已停止"}


def _yaml_to_py(case_file: str, project_id: str = "default") -> str:
    """将 case_file 映射为 pytest 脚本路径。
    默认项目：cases/api/test_x.yaml -> tests/api/test_x.py
    新项目：  cases/{pid}/api/test_x.yaml -> tests/{pid}/api/test_x.py
    找不到则返回空字符串。
    """
    if not case_file:
        return ""
    if project_id == "default":
        # 默认项目：cases/api/... -> tests/api/...
        rel = case_file.replace("cases/", "tests/", 1)
    else:
        # 新项目：cases/{pid}/api/... -> tests/{pid}/api/...
        rel = case_file.replace("cases/", "tests/", 1)
    if rel.endswith(".yaml"):
        rel = rel[:-5] + ".py"
    if not os.path.exists(os.path.join(settings.ROOT, rel)):
        return ""
    return rel


def _build_pytest_args(module: str, case_files: list, results_dir: str,
                       project_id: str = "default", report_dir: str = "") -> list:
    args = [sys.executable, "-m", "pytest", "-v", "--tb=short",
            "-W", "ignore::Warning", f"--alluredir={results_dir}",
            "-p", "no:cacheprovider"]
    # pytest-html 静态报告（飞书下载用，无需服务）
    # 放到 results_dir，避免被 allure generate --clean 删除
    if report_dir:
        args.append(f"--html={results_dir}/pytest_report.html")
        args.append("--self-contained-html")

    # 并行策略：不同 YAML 文件并行执行，同一文件内用例串行
    # --dist=loadfile: 按测试文件分组分配到不同 worker
    if module == "api":
        # 接口测试：完全并行（无首页限流风险）
        args.extend(["-n", "auto", "--dist=loadfile"])
    elif module == "ui":
        # UI 测试：限制 2 个并发（避免被测系统首页限流）
        args.extend(["-n", "2", "--dist=loadfile"])
    else:
        # 全部模块：串行执行
        args.extend(["-n", "1"])

    targets = []
    if case_files:
        for f in case_files:
            py = _yaml_to_py(f, project_id=project_id)
            if py:
                targets.append(py)
    if not targets:
        if case_files:
            # 指定了 YAML 但未找到对应 pytest 脚本，避免误跑全量 tests/
            return args
        if project_id == "default":
            if module == "ui":
                targets = ["tests/ui/"]
            elif module == "api":
                targets = ["tests/api/"]
            else:
                targets = ["tests/"]
        else:
            targets = [f"tests/{project_id}/"]
    args.extend(targets)
    if module == "api":
        args.extend(["-m", "api"])
    elif module == "ui":
        args.extend(["-m", "ui"])
    return args


def _parse_results(results_dir: str) -> dict:
    passed = failed = broken = total = 0
    for f in glob.glob(os.path.join(results_dir, "*-result.json")):
        try:
            with open(f, encoding="utf-8") as fp:
                st = json.load(fp).get("status", "")
            total += 1
            if st == "passed":
                passed += 1
            elif st == "failed":
                failed += 1
            elif st == "broken":
                broken += 1
        except Exception:
            pass
    return {"passed": passed, "failed": failed + broken, "total": total}


def _auto_save_defects(results_dir: str, task: dict):
    """解析失败用例，自动保存为缺陷到缺陷库。

    判断逻辑：
    - status=failed/broken 的用例视为缺陷
    - 已存在相同 task_id+tc_id 的缺陷不重复保存
    - 自动推断严重等级：broken>failed，含"超时"降为中
    - 同步写入 /Users/tanzsongsen/Documents/缺陷库.xlsx（含截图和日志）
    """
    import uuid as _uuid
    try:
        from backend import db as _db
    except Exception:
        return 0

    saved = 0
    now = _now()
    task_id = task.get("id", "")
    project_id = task.get("project_id", "default")

    # 读取任务完整日志（用于 xlsx 日志列）
    task_log = ""
    try:
        task_log = log_store.read(task_id) or ""
    except Exception:
        pass

    for f in glob.glob(os.path.join(results_dir, "*-result.json")):
        try:
            with open(f, encoding="utf-8") as fp:
                result = json.load(fp)
        except Exception:
            continue

        status = result.get("status", "")
        if status not in ("failed", "broken"):
            continue

        # 提取用例信息
        name = result.get("name", "unknown")
        full_name = result.get("fullName", "")  # 如 tests.ui.test_jdy_flow.test_smoke_full_flow[case0]
        labels = result.get("labels", [])
        tc_id = ""
        scenario = ""
        for label in labels:
            if label.get("name") == "tc_id":
                tc_id = label.get("value", "")
            elif label.get("name") == "scenario":
                scenario = label.get("value", "")

        # 从 fullName 解析测试文件路径和函数名
        # fullName 格式：tests.ui.test_jdy_flow.test_smoke_full_flow[case0]
        # 转换为：tests/ui/test_jdy_flow.py::test_smoke_full_flow[case0]
        test_file_path = ""
        test_func_name = name
        if full_name:
            parts = full_name.rsplit(".", 1)
            if len(parts) == 2:
                module_path = parts[0]  # tests.ui.test_jdy_flow
                test_func_name = parts[1]  # test_smoke_full_flow[case0]
                test_file_path = module_path.replace(".", "/") + ".py"

        # 从 parameters 中的 case 参数提取 scenario、tc_id 和 case_file_path（YAML 数据驱动场景）
        case_file_path = ""
        if not scenario or not tc_id:
            import ast as _ast
            for p in result.get("parameters", []):
                if p.get("name") == "case":
                    try:
                        case_dict = _ast.literal_eval(p.get("value", "{}"))
                        if isinstance(case_dict, dict):
                            if not scenario:
                                scenario = case_dict.get("scenario", "")
                            if not tc_id:
                                tc_id = case_dict.get("tc_id", "")
                            case_file_path = case_dict.get("case_file_path", "")
                    except Exception:
                        pass

        # 错误信息
        status_details = result.get("statusDetails", {})
        error_msg = status_details.get("message", "") or status_details.get("trace", "")
        # 截取前 2000 字符避免过长
        error_msg = error_msg[:2000] if error_msg else ""

        # 提取截图路径（allure attachments）
        screenshot_path = ""
        attachments = result.get("attachments", [])
        for att in attachments:
            if att.get("type", "").startswith("image/"):
                src = att.get("source", "")
                if src:
                    full_path = os.path.join(results_dir, src)
                    if os.path.exists(full_path):
                        screenshot_path = full_path
                        break
        # 兜底：从 screenshots/task_id/ 目录查找含用例名的截图
        if not screenshot_path:
            shots_dir = os.path.join(settings.SCREENSHOTS_DIR, task_id)
            if os.path.isdir(shots_dir):
                for shot in sorted(os.listdir(shots_dir), reverse=True):
                    if name in shot and shot.endswith(".png"):
                        screenshot_path = os.path.join(shots_dir, shot)
                        break

        # 提取该用例相关的日志片段（用例名到下一个用例名之间）
        case_log = ""
        if task_log and name:
            idx = task_log.find(name)
            if idx >= 0:
                # 往后取 2000 字符或到下一个测试用例
                segment = task_log[idx:idx + 3000]
                case_log = segment[:2000]
            else:
                case_log = error_msg or "(未找到用例日志)"
        else:
            case_log = error_msg or "(无日志)"

        # 推断严重等级
        severity = "中"
        if status == "broken":
            severity = "高"
        if error_msg and ("超时" in error_msg or "timeout" in error_msg.lower()):
            severity = "中"

        # 推断错误类型
        error_type = "断言失败" if status == "failed" else "执行异常"
        if error_msg:
            if "TimeoutError" in error_msg or "超时" in error_msg:
                error_type = "超时"
            elif "AssertionError" in error_msg or "assert" in error_msg.lower():
                error_type = "断言失败"
            elif "SyntaxError" in error_msg:
                error_type = "代码错误"
            elif "Operation not permitted" in error_msg:
                error_type = "环境错误"

        # 去重：相同 task_id+tc_id 不重复保存
        if tc_id:
            existing = _db.execute(
                "SELECT id FROM defects WHERE task_id=? AND tc_id=?",
                (task_id, tc_id), fetch=True
            )
            if existing:
                continue

        # 构造缺陷标题（scenario + 失败）
        scenario_text = scenario or name
        title = f"{scenario_text}失败"[:100]

        # 构造执行路径：YAML文件路径 > [tc_id] scenario
        # 方便查找是执行了哪个 case 导致的缺陷
        exec_parts = []
        if case_file_path:
            exec_parts.append(case_file_path)
        elif test_file_path:
            exec_parts.append(f"{test_file_path}::{test_func_name}")
        else:
            exec_parts.append(test_func_name or name)
        if tc_id:
            exec_parts.append(f"[{tc_id}] {scenario_text}")
        elif scenario_text:
            exec_parts.append(scenario_text)
        execution_path = " > ".join(exec_parts)[:300]

        did = _uuid.uuid4().hex[:12]
        _db.execute(
            "INSERT INTO defects(id,task_id,project_id,tc_id,scenario,title,severity,"
            "error_type,error_message,page_url,screenshot,source,status,created_at,updated_at) "
            "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (did, task_id, project_id, tc_id, scenario, title, severity,
             error_type, error_msg, execution_path, screenshot_path, "auto", "open", now, now),
        )
        saved += 1

    if saved:
        log_store.write(task_id, f"🐛 自动保存 {saved} 条缺陷到缺陷库\n")
        # 同步写入项目配置的 xlsx 路径
        try:
            from backend.projects import service as _proj_svc
            proj = _proj_svc.get_project(project_id)
            xlsx_path = proj.get("defect_xlsx_path", "") if proj else ""
            if xlsx_path:
                # 展开路径中的 ~ 符号
                xlsx_path = os.path.expanduser(xlsx_path)
                _sync_defects_to_xlsx(xlsx_path, project_id=project_id)
                log_store.write(task_id, f"📥 已同步缺陷到 {xlsx_path}\n")
            else:
                log_store.write(task_id, "⚠️ 项目未配置缺陷库xlsx路径，跳过同步\n")
        except Exception as e:
            log_store.write(task_id, f"⚠️ 同步缺陷到xlsx失败: {e}\n")
    return saved


def _sync_defects_to_xlsx(xlsx_path: str = "/Users/tanzsongsen/Documents/缺陷库.xlsx",
                          project_id: str = ""):
    """将缺陷库同步写入 xlsx 文件，包含截图和日志。

    xlsx 结构：标题 | 执行路径 | 严重等级 | 截图 | 日志
    截图：嵌入图片到单元格
    日志：错误信息 + 用例相关日志
    project_id：为空则导出全部缺陷，否则只导出该项目的缺陷
    """
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XlImage
        from backend import db as _db
    except Exception as e:
        raise RuntimeError(f"openpyxl 导入失败: {e}")

    # 读取缺陷（按 project_id 过滤），包含 page_url（存储执行路径）
    if project_id:
        rows = _db.execute(
            "SELECT title, severity, screenshot, error_message, scenario, tc_id, task_id, page_url "
            "FROM defects WHERE project_id=? ORDER BY created_at DESC",
            (project_id,), fetch=True
        )
    else:
        rows = _db.execute(
            "SELECT title, severity, screenshot, error_message, scenario, tc_id, task_id, page_url "
            "FROM defects ORDER BY created_at DESC", fetch=True
        )

    # 加载或创建 xlsx
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    # 使用固定 sheet 名"国信小米"（与现有文件保持一致）
    sheet_name = "国信小米"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 清空旧内容（保留表头）
        for row in list(ws.iter_rows(min_row=2)):
            for cell in row:
                cell.value = None
        # 清除旧图片
        for img in list(ws._images):
            ws._images.remove(img)
    else:
        ws = wb.create_sheet(sheet_name)

    # 写表头：标题 | 执行路径 | 严重等级 | 截图 | 日志
    headers = ["标题", "执行路径", "严重等级", "截图", "日志"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    # 设置列宽
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 95
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 60

    # 写数据
    for idx, row in enumerate(rows, 2):
        title = row["title"] if isinstance(row, dict) else row[0]
        severity = row["severity"] if isinstance(row, dict) else row[1]
        screenshot = row["screenshot"] if isinstance(row, dict) else row[2]
        error_msg = row["error_message"] if isinstance(row, dict) else row[3]
        scenario = row["scenario"] if isinstance(row, dict) else row[4]
        tc_id = row["tc_id"] if isinstance(row, dict) else row[5]
        name = row.get("task_id", "") if isinstance(row, dict) else row[6]
        # page_url 字段存储执行路径（测试文件::函数 > [tc_id] scenario）
        page_url = row.get("page_url", "") if isinstance(row, dict) else (row[7] if len(row) > 7 else "")

        # 标题列：scenario + 失败（已存于 title 字段）
        ws.cell(row=idx, column=1, value=str(title or ""))
        # 执行路径列：优先使用 page_url 中存储的完整执行路径
        if page_url:
            exec_path = str(page_url)[:300]
        else:
            # 回退：[tc_id] scenario
            exec_path = f"[{tc_id or name}] {scenario or name}"[:200]
        ws.cell(row=idx, column=2, value=exec_path)
        ws.cell(row=idx, column=3, value=str(severity or "中"))
        # 日志列：错误信息
        log_text = str(error_msg or "")
        if len(log_text) > 3000:
            log_text = log_text[:3000] + "...(截断)"
        ws.cell(row=idx, column=5, value=log_text)

        # 嵌入截图（D列）
        if screenshot and os.path.exists(screenshot):
            try:
                img = XlImage(screenshot)
                # 缩放图片到合理大小
                max_w, max_h = 300, 200
                if img.width > max_w:
                    ratio = max_w / img.width
                    img.width = max_w
                    img.height = int(img.height * ratio)
                if img.height > max_h:
                    ratio = max_h / img.height
                    img.height = max_h
                    img.width = int(img.width * ratio)
                img.anchor = f"D{idx}"
                ws.add_image(img)
            except Exception:
                ws.cell(row=idx, column=4, value=f"(截图无法嵌入: {screenshot})")
        else:
            ws.cell(row=idx, column=4, value="(无截图)")

        # 设置行高（适应图片）
        ws.row_dimensions[idx].height = 150

    wb.save(xlsx_path)
    return len(rows)


def _import_defects_from_xlsx(xlsx_path: str, project_id: str = ""):
    """从 xlsx 文件导入缺陷到数据库。

    xlsx 结构：标题 | 执行路径 | 严重等级 | 截图 | 日志
    与 _sync_defects_to_xlsx 互为逆操作。
    返回导入的条数。
    """
    if not xlsx_path or not os.path.exists(xlsx_path):
        raise RuntimeError(f"xlsx 文件不存在: {xlsx_path}")
    try:
        import openpyxl
        from backend import db as _db
    except Exception as e:
        raise RuntimeError(f"openpyxl 导入失败: {e}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    imported = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # 跳过表头
        for row in rows[1:]:
            if not row or not any(row):
                continue
            title = str(row[0] or "").strip() if len(row) > 0 else ""
            exec_path = str(row[1] or "").strip() if len(row) > 1 else ""
            severity = str(row[2] or "中").strip() if len(row) > 2 else "中"
            screenshot = str(row[3] or "").strip() if len(row) > 3 else ""
            # 截图列可能是图片占位文字，如果是 "(无截图)" 或 "(截图无法嵌入...)" 则置空
            if screenshot.startswith("(") and screenshot.endswith(")"):
                screenshot = ""
            log_text = str(row[4] or "").strip() if len(row) > 4 else ""

            if not title:
                continue

            # 从执行路径解析 tc_id 和 scenario
            # 新格式: tests/ui/test_jdy_flow.py::test_smoke_full_flow[case0] > [TC-XXX] scenario
            # 旧格式: [TC-XXX] scenario 描述
            tc_id = ""
            scenario = exec_path
            # 优先从 " > " 后面部分解析 [tc_id] scenario
            parse_part = exec_path
            if " > " in exec_path:
                parse_part = exec_path.rsplit(" > ", 1)[-1]
            if parse_part.startswith("["):
                end = parse_part.find("]")
                if end > 0:
                    tc_id = parse_part[1:end]
                    scenario = parse_part[end + 1:].strip()

            # 去重检查：同标题+同 tc_id 的缺陷不重复导入
            existing = _db.execute(
                "SELECT id FROM defects WHERE title=? AND tc_id=? AND project_id=?",
                (title, tc_id, project_id), fetch=True
            )
            if existing:
                continue

            import uuid as _uuid
            import time as _time
            did = _uuid.uuid4().hex[:12]
            now = _time.strftime("%Y-%m-%d %H:%M:%S")
            # page_url 存储完整执行路径，方便后续导出
            _db.execute(
                "INSERT INTO defects(id,task_id,project_id,tc_id,scenario,title,severity,"
                "error_type,error_message,page_url,screenshot,source,status,created_at,updated_at) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (did, "", project_id, tc_id, scenario, title, severity,
                 "imported", log_text, exec_path, screenshot, "xlsx_import", "open", now, now),
            )
            imported += 1
    wb.close()
    return imported


def _write_environment(results_dir: str, task: dict, stats: dict):
    env_name = task.get("env", "test")
    # 优先使用任务中存储的 base_url，其次从 config/{env}.yaml 读取
    base_url = task.get("base_url") or ""
    if not base_url:
        env_config_path = os.path.join(settings.ROOT, "config", f"{env_name}.yaml")
        if os.path.exists(env_config_path):
            try:
                with open(env_config_path, "r", encoding="utf-8") as f:
                    env_info = yaml.safe_load(f) or {}
                    base_url = env_info.get("base_url", "")
            except Exception:
                pass

    os.makedirs(results_dir, exist_ok=True)
    env_file = os.path.join(results_dir, "environment.properties")
    with open(env_file, "w", encoding="utf-8") as f:
        f.write(f"URL={base_url}\n")
        f.write(f"ENV={env_name}\n")
        f.write(f"PROJECT={task.get('project_id', '')}\n")


def _generate_allure(results_dir: str, report_dir: str) -> bool:
    try:
        subprocess.run(
            [settings.ALLURE_CLI, "generate", results_dir, "-o", report_dir, "--clean"],
            check=False, timeout=300,
        )

        env_src = os.path.join(results_dir, "environment.properties")
        env_dst = os.path.join(report_dir, "data", "environment.properties")
        if os.path.exists(env_src):
            os.makedirs(os.path.dirname(env_dst), exist_ok=True)
            shutil.copy2(env_src, env_dst)

        return os.path.exists(os.path.join(report_dir, "index.html"))
    except Exception:
        return False


def _generate_static_report(results_dir: str, report_dir: str):
    """从 allure-results 生成自包含的静态 report.html。

    所有 CSS/JS/数据内联到单个 HTML 文件中，
    下载后用浏览器直接打开即可查看，无需 HTTP 服务。
    """
    import html as _html

    # 读取所有 result 文件
    test_cases = []
    suites = {}
    for f in sorted(glob.glob(os.path.join(results_dir, "*-result.json"))):
        try:
            with open(f, encoding="utf-8") as fp:
                result = json.load(fp)
        except Exception:
            continue

        name = result.get("name", "unknown")
        status = result.get("status", "unknown")
        duration_ms = (result.get("stop", 0) or 0) - (result.get("start", 0) or 0)

        # 提取 suite / feature / story
        suite_name = ""
        feature = ""
        story = ""
        for label in result.get("labels", []):
            ln = label.get("name", "")
            lv = label.get("value", "")
            if ln == "suite":
                suite_name = lv
            elif ln == "feature":
                feature = lv
            elif ln == "story":
                story = lv
        if not suite_name:
            full_name = result.get("fullName", name)
            suite_name = full_name.rsplit(".", 1)[0] if "." in full_name else "default"

        # 提取 step 和 attachment 信息
        steps_info = []
        for step in result.get("steps", []):
            step_name = step.get("name", "")
            step_status = step.get("status", "")
            step_dur = (step.get("stop", 0) or 0) - (step.get("start", 0) or 0)
            step_attachments = []
            for att in step.get("attachments", []):
                att_name = att.get("name", "")
                att_source = att.get("source", "")
                # 读取 attachment 内容
                att_content = ""
                if att_source:
                    att_path = os.path.join(results_dir, att_source)
                    if os.path.exists(att_path):
                        try:
                            with open(att_path, encoding="utf-8") as af:
                                att_content = af.read()[:2000]
                        except Exception:
                            pass
                step_attachments.append({"name": att_name, "content": att_content})
            steps_info.append({
                "name": step_name, "status": step_status,
                "duration": step_dur, "attachments": step_attachments,
            })

        # 错误信息
        status_details = result.get("statusDetails", {})
        error_msg = status_details.get("message", "") or status_details.get("trace", "")
        if error_msg and len(error_msg) > 3000:
            error_msg = error_msg[:3000] + "..."

        # parameters
        params = result.get("parameters", [])

        test_cases.append({
            "name": name, "status": status, "duration_ms": duration_ms,
            "suite": suite_name, "feature": feature, "story": story,
            "steps": steps_info, "error": error_msg, "params": params,
        })

        if suite_name not in suites:
            suites[suite_name] = {"passed": 0, "failed": 0, "broken": 0, "skipped": 0, "total": 0}
        suites[suite_name]["total"] += 1
        if status in suites[suite_name]:
            suites[suite_name][status] += 1

    if not test_cases:
        return

    # 统计
    total = len(test_cases)
    passed = sum(1 for t in test_cases if t["status"] == "passed")
    failed = sum(1 for t in test_cases if t["status"] == "failed")
    broken = sum(1 for t in test_cases if t["status"] == "broken")
    skipped = sum(1 for t in test_cases if t["status"] == "skipped")
    total_duration = sum(t["duration_ms"] for t in test_cases)

    # 读取 environment.properties
    env_info = {}
    env_file = os.path.join(results_dir, "environment.properties")
    if os.path.exists(env_file):
        for line in open(env_file, encoding="utf-8"):
            if "=" in line:
                k, v = line.strip().split("=", 1)
                env_info[k] = v

    # 转义辅助
    def _e(s):
        return _html.escape(str(s))

    def _status_color(s):
        return {"passed": "#97cc64", "failed": "#fd5454", "broken": "#ffd05b",
                "skipped": "#aaa"}.get(s, "#aaa")

    def _status_icon(s):
        return {"passed": "✅", "failed": "❌", "broken": "⚠️", "skipped": "⏭️"}.get(s, "❓")

    # 构建 HTML
    html_parts = [
        '<!DOCTYPE html><html lang="zh-CN"><head><meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width,initial-scale=1">',
        '<title>测试报告</title>',
        '<style>',
        '*{margin:0;padding:0;box-sizing:border-box}body{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;background:#f5f5f5;color:#333;line-height:1.6}',
        '.container{max-width:1200px;margin:0 auto;padding:20px}',
        '.header{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:#fff;padding:30px;border-radius:12px;margin-bottom:20px}',
        '.header h1{font-size:24px;margin-bottom:10px}.stats{display:flex;gap:20px;flex-wrap:wrap}',
        '.stat{background:rgba(255,255,255,.2);padding:10px 20px;border-radius:8px;text-align:center;min-width:80px}',
        '.stat .num{font-size:28px;font-weight:bold}.stat .label{font-size:12px;opacity:.8}',
        '.env-bar{background:#fff;padding:12px 20px;border-radius:8px;margin-bottom:20px;display:flex;gap:20px;flex-wrap:wrap;font-size:13px;color:#666}',
        '.suite{background:#fff;border-radius:8px;margin-bottom:16px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,.06)}',
        '.suite-header{padding:14px 20px;cursor:pointer;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid #eee}',
        '.suite-header:hover{background:#fafafa}.suite-name{font-weight:bold;font-size:15px}',
        '.suite-stats{display:flex;gap:12px;font-size:13px}',
        '.suite-body{display:none}.suite.open .suite-body{display:block}',
        '.case{padding:12px 20px;border-bottom:1px solid #f0f0f0}.case:last-child{border:none}',
        '.case-header{display:flex;align-items:center;gap:8px;cursor:pointer}',
        '.case-header:hover .case-name{text-decoration:underline}',
        '.case-name{flex:1;font-size:14px}.case-dur{color:#999;font-size:12px;min-width:60px;text-align:right}',
        '.case-detail{display:none;margin:8px 0 0 28px;padding:10px;background:#f8f9fa;border-radius:6px;font-size:13px;white-space:pre-wrap;word-break:break-all}',
        '.case.open .case-detail{display:block}',
        '.step{margin:4px 0;padding:4px 8px;border-left:3px solid #ddd}.step-name{font-weight:500}',
        '.att{margin:2px 0 2px 16px;color:#666;font-size:12px}.att-content{background:#eee;padding:6px;border-radius:4px;margin:4px 0;max-height:200px;overflow:auto;white-space:pre-wrap;font-family:monospace;font-size:12px}',
        '.badge{display:inline-block;padding:2px 8px;border-radius:10px;color:#fff;font-size:11px;font-weight:bold}',
        '.progress{height:8px;border-radius:4px;overflow:hidden;display:flex;margin-top:10px}',
        '.progress div{height:100%}',
        '.filter-bar{margin-bottom:16px;display:flex;gap:8px;flex-wrap:wrap}',
        '.filter-btn{padding:6px 14px;border:1px solid #ddd;border-radius:6px;background:#fff;cursor:pointer;font-size:13px}',
        '.filter-btn.active{background:#667eea;color:#fff;border-color:#667eea}',
        '.filter-btn:hover{border-color:#667eea}',
        '</style></head><body><div class="container">',
        # Header
        '<div class="header">',
        '<h1>📊 测试执行报告</h1>',
        '<div class="stats">',
        f'<div class="stat"><div class="num">{total}</div><div class="label">总计</div></div>',
        f'<div class="stat"><div class="num" style="color:#97cc64">{passed}</div><div class="label">通过</div></div>',
        f'<div class="stat"><div class="num" style="color:#fd5454">{failed}</div><div class="label">失败</div></div>',
        f'<div class="stat"><div class="num" style="color:#ffd05b">{broken}</div><div class="label">异常</div></div>',
        f'<div class="stat"><div class="num" style="color:#aaa">{skipped}</div><div class="label">跳过</div></div>',
        f'<div class="stat"><div class="num">{total_duration/1000:.1f}s</div><div class="label">耗时</div></div>',
        '</div>',
        '<div class="progress">',
        f'<div style="width:{passed/total*100:.1f}%;background:#97cc64"></div>' if total else '',
        f'<div style="width:{failed/total*100:.1f}%;background:#fd5454"></div>' if total else '',
        f'<div style="width:{broken/total*100:.1f}%;background:#ffd05b"></div>' if total else '',
        '</div></div>',
    ]

    # Environment info
    if env_info:
        html_parts.append('<div class="env-bar">')
        for k, v in env_info.items():
            html_parts.append(f'<span><b>{_e(k)}</b>: {_e(v)}</span>')
        html_parts.append('</div>')

    # Filter bar
    html_parts.append('<div class="filter-bar">')
    html_parts.append('<button class="filter-btn active" onclick="filter(\'all\')">全部</button>')
    html_parts.append('<button class="filter-btn" onclick="filter(\'passed\')">✅ 通过</button>')
    html_parts.append('<button class="filter-btn" onclick="filter(\'failed\')">❌ 失败</button>')
    html_parts.append('<button class="filter-btn" onclick="filter(\'broken\')">⚠️ 异常</button>')
    html_parts.append('</div>')

    # Suites & cases
    for suite_name in sorted(suites.keys()):
        s = suites[suite_name]
        suite_cases = [t for t in test_cases if t["suite"] == suite_name]
        html_parts.append(f'<div class="suite open" data-suite="{_e(suite_name)}">')
        html_parts.append(
            f'<div class="suite-header" onclick="toggleSuite(this)">'
            f'<span class="suite-name">{_e(suite_name)}</span>'
            f'<span class="suite-stats">'
            f'<span class="badge" style="background:#97cc64">{s["passed"]} 通过</span> '
            f'<span class="badge" style="background:#fd5454">{s["failed"]} 失败</span> '
            f'<span class="badge" style="background:#ffd05b">{s["broken"]} 异常</span>'
            f'</span></div>'
        )
        html_parts.append('<div class="suite-body">')
        for tc in suite_cases:
            icon = _status_icon(tc["status"])
            color = _status_color(tc["status"])
            html_parts.append(
                f'<div class="case" data-status="{tc["status"]}">'
                f'<div class="case-header" onclick="toggleCase(this)">'
                f'<span>{icon}</span>'
                f'<span class="case-name">{_e(tc["name"])}</span>'
                f'<span class="case-dur">{tc["duration_ms"]:.0f}ms</span>'
                f'</div>'
                f'<div class="case-detail">'
            )
            # Feature / Story
            if tc["feature"]:
                html_parts.append(f'<b>Feature:</b> {_e(tc["feature"])}<br>')
            if tc["story"]:
                html_parts.append(f'<b>Story:</b> {_e(tc["story"])}<br>')
            # Error
            if tc["error"]:
                html_parts.append(f'<b style="color:#fd5454">错误信息:</b><br>{_e(tc["error"])}<br><br>')
            # Steps & attachments
            for step in tc["steps"]:
                step_icon = _status_icon(step["status"])
                html_parts.append(
                    f'<div class="step" style="border-left-color:{_status_color(step["status"])}">'
                    f'{step_icon} <span class="step-name">{_e(step["name"])}</span> '
                    f'<span style="color:#999;font-size:11px">({step["duration"]:.0f}ms)</span>'
                )
                for att in step["attachments"]:
                    html_parts.append(f'<div class="att">📎 {_e(att["name"])}')
                    if att["content"]:
                        html_parts.append(f'<div class="att-content">{_e(att["content"])}</div>')
                    html_parts.append('</div>')
                html_parts.append('</div>')
            html_parts.append('</div></div>')  # case-detail, case
        html_parts.append('</div></div>')  # suite-body, suite

    # JavaScript
    html_parts.append("""
<script>
function toggleSuite(el){el.parentElement.classList.toggle('open')}
function toggleCase(el){el.parentElement.classList.toggle('open')}
function filter(status){
    document.querySelectorAll('.filter-btn').forEach(b=>b.classList.remove('active'));
    event.target.classList.add('active');
    document.querySelectorAll('.case').forEach(c=>{
        c.style.display=(status==='all'||c.dataset.status===status)?'':'none';
    });
    document.querySelectorAll('.suite').forEach(s=>{
        const visible=s.querySelectorAll('.case[style=""],.case:not([style])');
        s.style.display=(status==='all'||s.querySelectorAll('.case[data-status="'+status+'"]').length>0)?'':'none';
    });
}
</script>""")
    html_parts.append('</div></body></html>')

    # 写入 report.html
    report_html_path = os.path.join(report_dir, "report.html")
    with open(report_html_path, "w", encoding="utf-8") as f:
        f.write("".join(html_parts))


def _update_task(task_id: str, **fields):
    cols = ", ".join(f"{k}=?" for k in fields)
    db.execute(f"UPDATE tasks SET {cols} WHERE id=?", tuple(fields.values()) + (task_id,))


def _get_task(task_id: str) -> dict:
    rows = db.execute(
        "SELECT t.*, p.name as project_name FROM tasks t "
        "LEFT JOIN projects p ON t.project_id=p.id WHERE t.id=?", (task_id,), fetch=True)
    return db.to_dict(rows[0]) if rows else {}


# 保存正在运行的子进程，用于停止功能
_running_procs: dict = {}


def run_task(task_id: str):
    """执行一个任务（由 API 直接调用或 worker 从队列消费后调用）。"""
    task = _get_task(task_id)
    if not task:
        return {"error": "task not found"}

    # 防止重复执行：如果任务已经是 running/success/failed 状态，跳过
    current_status = (task.get("status") or "").lower()
    if current_status in ("running", "success", "failed"):
        return {"error": f"task already {current_status}"}
    # stopped 任务也不重新执行（避免队列残留导致重复执行）
    if current_status == "stopped":
        return {"error": "task was stopped, not re-running"}

    module = task.get("module") or "all"
    case_files = json.loads(task.get("case_files") or "[]")
    base, results_dir, report_dir = _task_dirs(task_id)
    os.makedirs(results_dir, exist_ok=True)
    os.makedirs(report_dir, exist_ok=True)

    _update_task(task_id, status="running", started_at=_now())
    log_store.write(task_id, f"🚀 任务 {task_id} 开始 | 项目={task['project_id']} 模块={module}\n")
    log_store.write(task_id, f"📌 用例文件: {case_files or '全部'}\n")

    env = os.environ.copy()
    env["TASK_ID"] = task_id  # 供 conftest 截图按任务归档
    env["ENV"] = task.get("env", "test")  # 供 yaml_loader.load_config 选择环境配置
    env["PROJECT_ID"] = task["project_id"]  # 供 conftest/yaml_loader 使用项目级配置
    # 设置 BASE_URL 环境变量（执行时指定的 Base URL，覆盖页面对象中的默认值）
    base_url = task.get("base_url") or ""
    if base_url:
        env["BASE_URL"] = base_url
        log_store.write(task_id, f"🌐 Base URL: {base_url}\n")

    cmd = _build_pytest_args(module, case_files, results_dir, project_id=task["project_id"], report_dir=report_dir)
    log_store.write(task_id, f"$ {' '.join(cmd)}\n\n")

    start = time.time()
    try:
        import signal as _signal
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            cwd=settings.ROOT, env=env, text=True, bufsize=1,
            preexec_fn=os.setsid,  # 新建进程组，便于整组终止
        )
        _running_procs[task_id] = proc
        for line in iter(proc.stdout.readline, ""):
            log_store.write(task_id, line)
        proc.wait()
        rc = proc.returncode
    except Exception as e:
        log_store.write(task_id, f"❌ 执行异常: {e}\n")
        rc = -1
    finally:
        _running_procs.pop(task_id, None)

    duration = round(time.time() - start, 2)
    stats = _parse_results(results_dir)

    # 检查是否被用户手动停止
    current = _get_task(task_id)
    if current.get("status") == "stopped":
        log_store.write(task_id, f"\n{'=' * 40}\n⛔ 任务已被用户停止 | 耗时={duration}s\n")
        return current

    status = "success" if stats["failed"] == 0 and rc == 0 else "failed"
    finished_at = _now()

    task["finished_at"] = finished_at
    _write_environment(results_dir, task, stats)

    # 自动保存失败用例为缺陷
    if stats["failed"] > 0:
        try:
            _auto_save_defects(results_dir, task)
        except Exception as e:
            log_store.write(task_id, f"⚠️ 自动保存缺陷失败: {e}\n")
    report_ok = _generate_allure(results_dir, report_dir)

    # 把 pytest-html 报告从 results_dir 复制到 report_dir
    pytest_html_src = os.path.join(results_dir, "pytest_report.html")
    if os.path.exists(pytest_html_src):
        shutil.copy2(pytest_html_src, os.path.join(report_dir, "pytest_report.html"))

    # 独立生成静态报告，不影响主 Allure 报告
    if report_ok:
        try:
            _generate_static_report(results_dir, report_dir)
        except Exception as e:
            log_store.write(task_id, f"⚠️ 静态报告生成失败: {e}\n")
    
    if report_ok:
        # 使用相对路径，无论用户通过 localhost / 局域网IP / 域名访问都能打开
        report_url = f"/report/tasks/{task_id}/report/"
    else:
        report_url = ""

    _update_task(
        task_id,
        status=status,
        passed=stats["passed"],
        failed=stats["failed"],
        total=stats["total"],
        duration=duration,
        finished_at=finished_at,
        report_url=report_url,
    )

    final = _get_task(task_id)
    log_store.write(
        task_id,
        f"\n{'=' * 40}\n✅ 执行完成 | 状态={status} | 通过={stats['passed']} "
        f"失败={stats['failed']} 总计={stats['total']} | 耗时={duration}s\n"
    )
    if report_ok:
        log_store.write(task_id, f"📊 报告: {report_url}\n")
        # 离线报告下载链接（zip 打包，服务关闭后也可查看）
        host = settings.EXTERNAL_URL or f"http://localhost:{settings.SERVER_CFG.get('port', 8000)}"
        download_url = f"{host}/api/tasks/{task_id}/report/download"
        log_store.write(task_id, f"📥 离线报告下载: {download_url}\n")

    # 飞书通知：飞书是外部应用，必须用绝对路径
    feishu_card_data = dict(final) if final else {}
    if report_ok:
        host = settings.EXTERNAL_URL or f"http://localhost:{settings.SERVER_CFG.get('port', 8000)}"
        feishu_card_data["report_url"] = f"{host}/report/tasks/{task_id}/report/"
        # 下载链接：下载 report.html，双击即可查看，无需服务
        feishu_card_data["report_download_url"] = f"{host}/api/tasks/{task_id}/report/download"
    feishu.send_card(feishu_card_data)
    return final


def create_task(project_id: str, module: str = "all", case_files: list = None,
                triggered_by: str = "manual", env: str = "test",
                base_url: str = "") -> str:
    task_id = uuid.uuid4().hex[:12]
    db.execute(
        "INSERT INTO tasks(id,project_id,module,case_files,env,status,triggered_by,started_at,base_url) "
        "VALUES(?,?,?,?,?,?,?,?,?)",
        (task_id, project_id, module, json.dumps(case_files or [], ensure_ascii=False),
         env, "pending", triggered_by, _now(), base_url),
    )
    return task_id


def dispatch(task_id: str) -> bool:
    """负载均衡调度：Redis 可用则入队由 worker 消费，否则直接本地执行。"""
    from backend import redis_queue
    if redis_queue.available():
        redis_queue.push_task({"task_id": task_id})
        return False  # 异步
    run_task(task_id)
    return True  # 同步已完成
